#! python3
# spreadsheetCellInverter.py - Inverts the rows and columns of spreadsheet.
# Using with formulae in data may have unintended results.

import openpyxl, sys

if len(sys.argv) == 2:

	try:
		location = str(sys.argv[1])
	
	except Exception as e:
		print(e)

	wb = openpyxl.load_workbook(location)
	
	before_sheet = wb.active
	before_sheet.title = 'Before'
	
	after_sheet = wb.create_sheet(index = 1, title = 'After')
	
	# x refers to column, and y refers to row.
	for x in range(1, before_sheet.max_column + 1):
		for y in range(1, before_sheet.max_row + 1):
			after_sheet.cell(row = x, column = y).value = before_sheet.cell(row = y, column = x).value
			
	p = location[:-5] + '_inverted.xlsx'

	wb.save(p)
	
	print('Spreadsheet successfully inverted and saved to ' + p)

else:
	print("You must include a filepath in your argument.")